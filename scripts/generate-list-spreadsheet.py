#!/usr/bin/env python3
"""Generate an xlsx + csv of all Tis the Podcast episodes with TMDB poster URLs."""

import csv
import re
import time
import xml.etree.ElementTree as ET

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill

FEED_URL = "https://feed.podbean.com/tisthepodcast/feed.xml"
OUT_XLSX = "/Users/thomcrowe/Documents/GitHub/tisthepodcast/episode-list.xlsx"
OUT_CSV  = "/Users/thomcrowe/Documents/GitHub/tisthepodcast/episode-list.csv"

TMDB_TOKEN = "eyJhbGciOiJIUzI1NiJ9.eyJhdWQiOiI4M2QzYTU2YWJhYzU3MzVhNzlmODdiYzY4MzUzMTM4NCIsIm5iZiI6MTc3NjAyODUzNy4xMzMsInN1YiI6IjY5ZGMwYjc5OTFjYTAxYjI5YTAyNzdhNiIsInNjb3BlcyI6WyJhcGlfcmVhZCJdLCJ2ZXJzaW9uIjoxfQ.i2tv_Gcm2iqOXpHF7cFlDuapm2d6DfgnYwTWbvBi-TU"
TMDB_HEADERS = {"Authorization": f"Bearer {TMDB_TOKEN}", "accept": "application/json"}
TMDB_IMG_BASE = "https://image.tmdb.org/t/p/w500"

NS = {
    "itunes": "http://www.itunes.com/dtds/podcast-1.0.dtd",
    "content": "http://purl.org/rss/1.0/modules/content/",
}


def extract_show_title(raw_title):
    match = re.search(r"\(([^)]+?)\)?\.?\s*$", raw_title)
    if match:
        return match.group(1).replace(r"\.{2,}$", "").strip()
    return raw_title.strip()


def get_tmdb_poster(title):
    """Search TMDB for a movie first, then TV. Returns full image URL or empty string."""
    try:
        for media_type in ("movie", "tv"):
            url = f"https://api.themoviedb.org/3/search/{media_type}"
            r = requests.get(url, headers=TMDB_HEADERS, params={"query": title, "page": 1}, timeout=10)
            if r.ok:
                results = r.json().get("results", [])
                if results and results[0].get("poster_path"):
                    return TMDB_IMG_BASE + results[0]["poster_path"]
    except Exception:
        pass
    return ""


print("Fetching RSS feed…")
r = requests.get(FEED_URL, timeout=30)
r.raise_for_status()

root = ET.fromstring(r.content)
channel = root.find("channel")
items = channel.findall("item")

episodes = []
for item in items:
    ep_num_el = item.find("itunes:episode", NS)
    if ep_num_el is None or not ep_num_el.text:
        continue
    raw_title = (item.findtext("title") or "").strip()
    link = (item.findtext("link") or "").strip()
    ep_id = link.replace("https://tisthepodcast.podbean.com/e/", "").rstrip("/")
    episodes.append({
        "ep": int(ep_num_el.text),
        "raw_title": raw_title,
        "show_title": extract_show_title(raw_title),
        "link": link,
        "id": ep_id,
    })

episodes.sort(key=lambda e: e["ep"])
print(f"Found {len(episodes)} numbered episodes. Fetching TMDB posters…")

for i, ep in enumerate(episodes):
    poster = get_tmdb_poster(ep["show_title"])
    ep["poster_url"] = poster
    status = "found" if poster else "none"
    print(f"  [{i+1}/{len(episodes)}] Ep {ep['ep']:>3}: {ep['show_title'][:40]:<40} — {status}")
    time.sleep(0.1)

# ── XLSX ──────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Sheet 1: Full episode list ────────────────────────────────────────────────
ws = wb.active
ws.title = "Episode List"

headers = ["Episode #", "Raw Title", "Show Title", "Poster URL", "Needs Poster?", "Episode Link"]
header_font  = Font(name="Arial", bold=True, color="FFFFFF")
header_fill  = PatternFill("solid", start_color="2E7D32")
missing_fill = PatternFill("solid", start_color="FFFF00")  # yellow highlight

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, ep in enumerate(episodes, 2):
    missing = not ep["poster_url"]
    ws.cell(row=row_idx, column=1, value=ep["ep"]).font = Font(name="Arial")
    ws.cell(row=row_idx, column=2, value=ep["raw_title"]).font = Font(name="Arial")
    ws.cell(row=row_idx, column=3, value=ep["show_title"]).font = Font(name="Arial")
    url_cell = ws.cell(row=row_idx, column=4, value=ep["poster_url"])
    url_cell.font = Font(name="Arial", color="0000FF", underline="single") if ep["poster_url"] else Font(name="Arial")
    needs_cell = ws.cell(row=row_idx, column=5, value="⚠ MISSING" if missing else "✓")
    needs_cell.font = Font(name="Arial", bold=missing, color="FF0000" if missing else "2E7D32")
    if missing:
        for c in range(1, 7):
            ws.cell(row=row_idx, column=c).fill = missing_fill
    ws.cell(row=row_idx, column=6, value=ep["link"]).font = Font(name="Arial")

ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 45
ws.column_dimensions["D"].width = 80
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 60
ws.freeze_panes = "A2"

# ── Sheet 2: Missing posters only ─────────────────────────────────────────────
ws2 = wb.create_sheet("Needs Poster")
missing_episodes = [ep for ep in episodes if not ep["poster_url"]]

ms_headers = ["Episode #", "Show Title", "Poster URL (fill in)", "TMDB Search Link"]
for col, header in enumerate(ms_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = PatternFill("solid", start_color="C62828")  # red header for attention
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, ep in enumerate(missing_episodes, 2):
    tmdb_search = f"https://www.themoviedb.org/search?query={requests.utils.quote(ep['show_title'])}"
    ws2.cell(row=row_idx, column=1, value=ep["ep"]).font = Font(name="Arial")
    ws2.cell(row=row_idx, column=2, value=ep["show_title"]).font = Font(name="Arial", bold=True)
    ws2.cell(row=row_idx, column=3, value="").fill = PatternFill("solid", start_color="FFF9C4")
    link_cell = ws2.cell(row=row_idx, column=4, value=tmdb_search)
    link_cell.font = Font(name="Arial", color="0000FF", underline="single")

ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 45
ws2.column_dimensions["C"].width = 80
ws2.column_dimensions["D"].width = 80
ws2.freeze_panes = "A2"

wb.save(OUT_XLSX)

# ── CSV ───────────────────────────────────────────────────────────────────────
with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for ep in episodes:
        missing = not ep["poster_url"]
        writer.writerow([ep["ep"], ep["raw_title"], ep["show_title"], ep["poster_url"],
                         "MISSING" if missing else "OK", ep["link"]])

print(f"\nDone! Wrote:\n  {OUT_XLSX}\n  {OUT_CSV}")
